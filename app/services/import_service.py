"""Import service — baca Excel dan masukkan data penerima ke DB.

Kolom template (row 1 = header, row 2 = petunjuk opsional, row 3+ = data):
  NIK | Nama Lengkap | Tempat Lahir | Tanggal Lahir | Jenis Kelamin |
  Alamat | No Telp | Penghasilan | Jumlah Tanggungan | Status Rumah | Nama Kategori

- Row 2 dengan NIK berisi kata "wajib" / "opsional" / "format" otomatis diskip.
- Baris kosong dilewati.
- NIK duplikat: dicatat sebagai skipped.
- Error per baris dicatat, baris lain tetap diproses.
- Fraud detection + priority scoring otomatis per baris.
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date
from typing import List, Optional

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import Kategori, Penerima, PenerimaKategori
from app.services.fraud_service import apply_fraud_findings, detect_fraud
from app.services.priority_service import update_priority


# ── Hasil import ─────────────────────────────────────────────────────────────

@dataclass
class RowError:
    row: int
    nik: str
    message: str


@dataclass
class ImportResult:
    total_rows: int = 0
    imported: int = 0
    skipped_duplicate: int = 0
    errors: List[RowError] = field(default_factory=list)

    @property
    def failed(self) -> int:
        return len(self.errors)


# ── Helper parsing ────────────────────────────────────────────────────────────

def _str(v) -> str:
    return str(v).strip() if v is not None else ""


def _float(v) -> float:
    try:
        return float(str(v).replace(",", ".").strip())
    except (ValueError, TypeError):
        return 0.0


def _int(v) -> int:
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return 0


def _date(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            from datetime import datetime as dt
            return dt.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _resolve_kategori(db: Session, nama_str: str) -> List[int]:
    """Ubah 'Fakir Miskin, Disabilitas' → list id kategori."""
    ids: List[int] = []
    if not nama_str:
        return ids
    for nama in nama_str.split(","):
        nama = nama.strip()
        if not nama:
            continue
        k = db.query(Kategori).filter(
            Kategori.nama.ilike(nama), Kategori.is_active.is_(True)
        ).first()
        if k:
            ids.append(k.id)
    return ids


# ── Fungsi utama ──────────────────────────────────────────────────────────────

def import_penerima_xlsx(db: Session, file_bytes: bytes) -> ImportResult:
    result = ImportResult()

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    result.total_rows = sum(1 for r in rows if any(c is not None for c in r))

    _PETUNJUK_HINTS = {"wajib", "opsional", "format", "contoh", "optional", "required"}

    for row_idx, row in enumerate(rows, start=2):
        # Lewati baris kosong
        if not any(c is not None for c in row):
            continue

        # Ambil kolom (toleran terhadap file pendek)
        def col(i):
            return row[i] if i < len(row) else None

        nik             = _str(col(0))

        # Lewati baris petunjuk dari template (bukan data nyata)
        if any(h in nik.lower() for h in _PETUNJUK_HINTS):
            continue

        nama            = _str(col(1))
        tempat_lahir    = _str(col(2)) or None
        tanggal_lahir   = _date(col(3))
        jenis_kelamin   = _str(col(4)) or None
        alamat          = _str(col(5))
        no_telp         = _str(col(6)) or None
        penghasilan     = _float(col(7))
        jml_tanggungan  = _int(col(8))
        status_rumah    = _str(col(9)) or None
        nama_kategori   = _str(col(10))

        # Validasi wajib
        if not nik or not nama or not alamat:
            result.errors.append(RowError(row_idx, nik or "-", "NIK, Nama, dan Alamat wajib diisi"))
            continue

        if len(nik) != 16 or not nik.isdigit():
            result.errors.append(RowError(row_idx, nik, "NIK harus 16 digit angka"))
            continue

        # Cek duplikat NIK
        if db.query(Penerima).filter(Penerima.nik == nik).first():
            result.skipped_duplicate += 1
            continue

        # Resolve kategori
        try:
            kategori_ids = _resolve_kategori(db, nama_kategori)
            kategori_objs = (
                db.query(Kategori)
                .filter(Kategori.id.in_(kategori_ids), Kategori.is_active.is_(True))
                .all()
                if kategori_ids else []
            )
        except Exception as exc:
            result.errors.append(RowError(row_idx, nik, f"Kategori error: {exc}"))
            continue

        try:
            p = Penerima(
                nik=nik,
                nama=nama,
                tempat_lahir=tempat_lahir,
                tanggal_lahir=tanggal_lahir,
                jenis_kelamin=jenis_kelamin,
                alamat=alamat,
                no_telp=no_telp,
                penghasilan=penghasilan,
                jumlah_tanggungan=jml_tanggungan,
                status_rumah=status_rumah,
            )
            db.add(p)
            db.flush()

            for k in kategori_objs:
                db.add(PenerimaKategori(penerima_id=p.id, kategori_id=k.id))
            db.flush()
            db.refresh(p)

            update_priority(p, db)

            findings = detect_fraud(db, p)
            apply_fraud_findings(db, p, findings)

            db.commit()
            result.imported += 1

        except Exception as exc:
            db.rollback()
            result.errors.append(RowError(row_idx, nik, str(exc)))

    wb.close()
    return result
