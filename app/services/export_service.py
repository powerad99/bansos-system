"""Export service: penerima & distribusi -> Excel, plus template import.

Pakai openpyxl langsung (bukan pandas) supaya formatting header bisa custom.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models import Distribusi, Penerima
from app.models.distribusi_bulanan import DistribusiBulanan
from app.models.distribusi_sosial import DistribusiSosial


_HEADER_FILL    = PatternFill(start_color="2E5077", end_color="2E5077", fill_type="solid")
_HEADER_FONT    = Font(bold=True, color="FFFFFF")
_NOTE_FILL      = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
_NOTE_FONT      = Font(italic=True, color="856404")
_EXAMPLE_FILL   = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")


def _write_header(ws, headers):
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize(ws):
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)


# ── Export Penerima ───────────────────────────────────────────────────────────

def export_penerima_xlsx(items: Iterable[Penerima]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Penerima"

    headers = [
        "ID", "NIK", "Nama", "Tempat Lahir", "Tanggal Lahir", "Jenis Kelamin",
        "Alamat", "No Telp", "Penghasilan", "Tanggungan", "Status Rumah",
        "Kategori", "Priority Score", "Priority Level",
        "Fraud Flag", "Fraud Reason", "Aktif", "Created At",
    ]
    _write_header(ws, headers)

    for p in items:
        kategori_str = ", ".join(k.nama for k in p.kategori_list) if p.kategori_assoc else ""
        ws.append([
            p.id,
            p.nik,
            p.nama,
            p.tempat_lahir or "",
            p.tanggal_lahir.strftime("%Y-%m-%d") if p.tanggal_lahir else "",
            p.jenis_kelamin or "",
            p.alamat,
            p.no_telp or "",
            p.penghasilan,
            p.jumlah_tanggungan,
            p.status_rumah or "",
            kategori_str,
            round(p.priority_score, 4),
            p.priority_level.value,
            "YA" if p.fraud_flag else "TIDAK",
            p.fraud_reason or "",
            "AKTIF" if p.is_active else "NON-AKTIF",
            p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else "",
        ])

    _autosize(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Export Distribusi ─────────────────────────────────────────────────────────

def export_distribusi_xlsx(items: Iterable[Distribusi]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Distribusi"

    headers = [
        "No Seri", "Penerima ID", "NIK Penerima", "Nama Penerima",
        "Jenis Bantuan", "Nominal", "Status", "Petugas", "Tanggal",
    ]
    _write_header(ws, headers)

    for d in items:
        ws.append([
            d.no_seri,
            d.penerima_id,
            d.penerima.nik if d.penerima else "",
            d.penerima.nama if d.penerima else "",
            d.jenis_bantuan,
            d.nominal,
            d.status.value,
            d.petugas.username if d.petugas else "",
            d.tanggal_distribusi.strftime("%Y-%m-%d %H:%M"),
        ])

    _autosize(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Export Distribusi Bulanan ─────────────────────────────────────────────────

_NAMA_BULAN = ["", "Januari", "Februari", "Maret", "April", "Mei", "Juni",
               "Juli", "Agustus", "September", "Oktober", "November", "Desember"]


def export_distribusi_bulanan_xlsx(items: Iterable[DistribusiBulanan]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Distribusi Bulanan"

    headers = [
        "ID", "Kode Seri", "NIK", "Nama Penerima", "Alamat",
        "Bulan", "Nama Bulan", "Tahun", "Status",
        "Petugas Konfirmasi", "Tanggal Konfirmasi", "Dibuat",
    ]
    _write_header(ws, headers)

    for d in items:
        nama_bulan = _NAMA_BULAN[d.bulan] if 1 <= d.bulan <= 12 else str(d.bulan)
        ws.append([
            d.id,
            d.penerima.kode_seri if d.penerima else "",
            d.penerima.nik if d.penerima else "",
            d.penerima.nama if d.penerima else "",
            d.penerima.alamat if d.penerima else "",
            d.bulan,
            nama_bulan,
            d.tahun,
            d.status.value,
            d.confirmed_by.full_name if d.confirmed_by else "",
            d.confirmed_at.strftime("%Y-%m-%d %H:%M") if d.confirmed_at else "",
            d.created_at.strftime("%Y-%m-%d %H:%M") if d.created_at else "",
        ])

    _autosize(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()



# ── Export Distribusi Sosial ──────────────────────────────────────────────────

def export_distribusi_sosial_xlsx(items: Iterable[DistribusiSosial]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Distribusi Sosial"

    headers = [
        "ID", "Kode Seri", "NIK", "Nama Penerima", "Alamat",
        "Bulan", "Nama Bulan", "Tahun", "Status",
        "Petugas Konfirmasi", "Tanggal Konfirmasi", "Dibuat",
    ]
    _write_header(ws, headers)

    for d in items:
        nama_bulan = _NAMA_BULAN[d.bulan] if 1 <= d.bulan <= 12 else str(d.bulan)
        ws.append([
            d.id,
            d.penerima.kode_seri if d.penerima else "",
            d.penerima.nik if d.penerima else "",
            d.penerima.nama if d.penerima else "",
            d.penerima.alamat if d.penerima else "",
            d.bulan,
            nama_bulan,
            d.tahun,
            d.status.value,
            d.confirmed_by.full_name if d.confirmed_by else "",
            d.confirmed_at.strftime("%Y-%m-%d %H:%M") if d.confirmed_at else "",
            d.created_at.strftime("%Y-%m-%d %H:%M") if d.created_at else "",
        ])

    _autosize(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Template Import Penerima ──────────────────────────────────────────────────

_IMPORT_HEADERS = [
    "NIK",
    "Nama Lengkap",
    "Tempat Lahir",
    "Tanggal Lahir",
    "Jenis Kelamin",
    "Alamat",
    "No Telp",
    "Penghasilan",
    "Jumlah Tanggungan",
    "Status Rumah",
    "Nama Kategori",
]

_IMPORT_NOTES = [
    "Wajib, 16 digit angka",
    "Wajib",
    "Opsional",
    "Opsional, format YYYY-MM-DD",
    "Opsional, L atau P",
    "Wajib",
    "Opsional",
    "Opsional, angka (contoh: 500000)",
    "Opsional, angka (contoh: 3)",
    "Opsional: milik / sewa / menumpang",
    "Opsional, pisah koma (contoh: Fakir Miskin,Disabilitas)",
]

_IMPORT_EXAMPLE = [
    "3201234567890001",
    "Budi Santoso",
    "Jakarta",
    "1985-06-17",
    "L",
    "Jl. Merdeka No. 10, Cibinong",
    "08123456789",
    "750000",
    "4",
    "sewa",
    "Fakir Miskin",
]


def get_penerima_import_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Template Import Penerima"
    ws.freeze_panes = "A3"

    # Row 1: Header
    _write_header(ws, _IMPORT_HEADERS)
    ws.row_dimensions[1].height = 22

    # Row 2: Petunjuk
    for i, note in enumerate(_IMPORT_NOTES, start=1):
        cell = ws.cell(row=2, column=i, value=note)
        cell.fill = _NOTE_FILL
        cell.font = _NOTE_FONT
        cell.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 30

    # Row 3: Contoh data
    for i, val in enumerate(_IMPORT_EXAMPLE, start=1):
        cell = ws.cell(row=3, column=i, value=val)
        cell.fill = _EXAMPLE_FILL
    ws.row_dimensions[3].height = 18

    _autosize(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_filename(prefix: str) -> str:
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{ts}.xlsx"
